from dbase import Database
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import  PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from xlsxwriter.workbook import Workbook
from xlsxwriter import exceptions
import pandas as pd
import os
import tkinter as tk


class pricesheet:
    def __init__(self):

        self.db = Database('Database/building_db.db')

    def table_output(self):
        self.db.drop_table_works()
        conn = sqlite3.connect('Database/building_db.db')
        cur = conn.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS results_rc (id INTEGER PRIMARY KEY, name text, exc DOUBLE, strfill DOUBLE, foundation DOUBLE, lean_conc DOUBLE,conc_wall DOUBLE,backfill DOUBLE, rc_column DOUBLE, rc_beam DOUBLE, concrete_slab DOUBLE, ground_slab DOUBLE, formwork DOUBLE, rebar DOUBLE, wire_mesh DOUBLE, rainwater_gutter DOUBLE, rainwater_pipe DOUBLE, insulation_sika DOUBLE, insulation_membrane DOUBLE, pe_sheet DOUBLE,inside_brick_wall DOUBLE,gypsum_part_wall DOUBLE,aliminum_suspended_ceiling DOUBLE,acoustical_ceiling DOUBLE,gypsum_suspended_ceiling DOUBLE,rockwool_suspended_ceiling DOUBLE ,non_slip_ceramic_tile DOUBLE ,glazed_ceramic_tile DOUBLE ,glazed_ceramic_skirting DOUBLE ,epoxy_painting_floor DOUBLE ,resistant_floor DOUBLE ,acid_tile DOUBLE ,raised_floor DOUBLE ,laminated_parquet DOUBLE ,int_plaster DOUBLE ,ceiling_plaster DOUBLE ,ceramic_wall_tile DOUBLE ,int_paint DOUBLE ,acid_resistant_int_paint DOUBLE ,ceiling_paint DOUBLE ,epoxy_wall_paint DOUBLE,pvc DOUBLE,water_stop DOUBLE,marble_cladding DOUBLE,alucobond_cladding DOUBLE)")

        cur.execute(
            "CREATE TABLE IF NOT EXISTS results_steel (id INTEGER PRIMARY KEY, name text,exc DOUBLE,strfill DOUBLE,foundation DOUBLE,lean_conc DOUBLE,conc_wall DOUBLE,backfill DOUBLE,steel_column DOUBLE,ground_slab DOUBLE,steel_formwork DOUBLE,steel_rebar DOUBLE,wire_mesh DOUBLE,steelweight DOUBLE,grouting DOUBLE,embedded_steel DOUBLE,anchorbolt_m24 DOUBLE,anchorbolt_m30 DOUBLE,insulation_sika DOUBLE,insulation_membrane DOUBLE,pe_sheet DOUBLE,handrail DOUBLE,steel_grating DOUBLE,steel_ladder DOUBLE,chequered_plate DOUBLE,water_stop DOUBLE, rainwater_gutter DOUBLE, rainwater_pipe DOUBLE,inside_brick_wall DOUBLE,gypsum_part_wall DOUBLE,aliminum_suspended_ceiling DOUBLE,acoustical_ceiling DOUBLE,gypsum_suspended_ceiling DOUBLE,rockwool_suspended_ceiling DOUBLE ,non_slip_ceramic_tile DOUBLE ,glazed_ceramic_tile DOUBLE ,glazed_ceramic_skirting DOUBLE ,epoxy_painting_floor DOUBLE ,resistant_floor DOUBLE ,acid_tile DOUBLE ,raised_floor DOUBLE ,laminated_parquet DOUBLE ,int_plaster DOUBLE ,ceiling_plaster DOUBLE ,ceramic_wall_tile DOUBLE ,int_paint DOUBLE ,acid_resistant_int_paint DOUBLE ,ceiling_paint DOUBLE ,epoxy_wall_paint DOUBLE,pvc DOUBLE)")

        cur.execute(
            "DELETE FROM  results_rc")
        cur.execute(
            "DELETE FROM results_steel")

        dfs = pd.read_excel('Price_Sheet/Price_Sheet.xlsx', sheet_name=None,engine='openpyxl')
        for table, df in dfs.items():
            df.to_sql(table, conn)

    def print(self,project_name):
        global filename_output
        for version in range(0, 1500):
            if os.path.isfile('Output/{}_BUILDING_TOOL_OUTPUT{}.xlsx'.format(project_name,version)):
                print("File exist")
            else:
                print("File not exist")
                filename_output = 'Output/{}_BUILDING_TOOL_OUTPUT{}.xlsx'.format(project_name,version)
                break

        workbook = Workbook(filename_output)
        worksheet = workbook.add_worksheet('BOQ')
        worksheet_rc = workbook.add_worksheet('RC_BUILDING_DETAIL')
        worksheet_steel = workbook.add_worksheet('STEEL_BUILDING_DETAIL')
        wrap_format = workbook.add_format(
            {'bold': True, 'font_color': 'black', 'align': 'center', 'border': 2, 'text_wrap': True,
             'valign': 'vcenter'})

        conn = sqlite3.connect('Database/building_db.db')
        c = conn.cursor()

        c.execute("select * from works")
        mysel = c.execute("select * from works")
        headers = []
        for k, desc in enumerate(mysel.description):
            worksheet.write(0, k, desc[0], wrap_format)
            headers.append(desc[0])

        for i, row in enumerate(mysel):
            for j, value in enumerate(row):
                worksheet.write(i + 2, j, value)
        print(headers)

        building_info = c.execute("SELECT width,length,storeyno,storeyheight FROM buildings")
        for k, desc in enumerate(building_info):
            explanation = "Width = {}m\nLength = {}m\nStoreyno = {}\nStoreyheight = {}m".format(desc[0],desc[1],desc[2],desc[3])
            worksheet.write(1, k+4, explanation, wrap_format)

        rc = c.execute("SELECT name,storeyno,storeyheight,structuraltype,insulationtype,soilproperty,beam_connection,ground_wall,basement,width,length, exc_depth, strfill_depth, foundation_depth, foundation_D1, foundation_D2, tie_beam_d1, tie_beam_d2,rc_column_width,rc_column_length,rc_column_rangex,rc_column_rangey,basement_wall_height,basement_wall_thickness,ground_wall_height,ground_wall_thickness,rc_beam_width,rc_beam_length,rc_secondary_beam_D1,rc_secondary_beam_D2,concrete_slab,ground_slab,paraphet_height,paraphet_thickness  FROM buildings WHERE structuraltype = 'RC'")
        rc_param = []
        for k2,desc2 in enumerate(rc.description):
            worksheet_rc.write(0, k2, desc2[0], wrap_format)
            rc_param.append(desc2[0])

        for i, row in enumerate(rc):
            for j, value in enumerate(row):
                worksheet_rc.write(i + 1, j, value)

        steel = c.execute("SELECT name,storeyno,storeyheight,structuraltype,steeltype,insulationtype,seperatedroom,middlefloor,soilproperty,beam_connection,ground_wall,basement,width,length, exc_depth,strfill_depth, foundation_depth, foundation_D1, foundation_D2, tie_beam_d1, tie_beam_d2,steel_pedestal_d1,steel_pedestal_d2,steel_pedestal_depth,steel_pedestal_rangex, steel_pedestal_rangey,ground_wall_height,ground_wall_thickness,steel_weight,grouting_depth,concrete_slab,ground_slab FROM buildings WHERE structuraltype = 'STEEL'")
        steel_param = []
        for k3,desc3 in enumerate(steel.description):
            worksheet_steel.write(0, k3, desc3[0], wrap_format)
            steel_param.append(desc3[0])

        for i, row in enumerate(steel):
            for j, value in enumerate(row):
                worksheet_steel.write(i + 1, j, value)


        while True:
            try:
                workbook.close()
            except xlsxwriter.exceptions.FileCreateError as e:
                MsgBox = tk.messagebox.showwarning('WARNING', 'PLEASE CLOSE {}'.format(filename_output))
                if MsgBox == 'Ok':
                    continue

            break
        # MODIFY
        wb = load_workbook(filename_output)
        ws = wb[wb.sheetnames[0]]
        ws2 = wb[wb.sheetnames[1]]
        ws3 = wb[wb.sheetnames[2]]

        Border1 = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                         right=Side(border_style='thin'))
        Border2 = Border(left=Side(border_style='thick'), bottom=Side(border_style='thick'),
                         top=Side(border_style='thick'),
                         right=Side(border_style='thick'))
        Font1 = Font(name='Times New Roman', size=11, bold=True)
        Font2 = Font(name='Times New Roman', size=11, bold=False)
        Alignment1 = Alignment(vertical="center", wrap_text=True, horizontal="left")
        Alignment2 = Alignment(vertical="center", wrap_text=True, horizontal="center")
        rotate = Alignment(textRotation=90, vertical="center", wrap_text=True, horizontal="center")
        grayfill = PatternFill(start_color='00C0C0C0',
                               end_color='00C0C0C0',
                               fill_type='solid')

        orangefill = PatternFill(start_color='00FF9900',
                                 end_color='00FF9900',
                                 fill_type='solid')

        bluefill = PatternFill(start_color='d6e5f2',
                                 end_color='d6e5f2',
                                 fill_type='solid')

        columns = ["C"]

        for column in columns:
            ws.column_dimensions[column].width = 40

        ws2.column_dimensions["A"].width = 40
        ws3.column_dimensions["A"].width = 40

        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=3).alignment = Alignment1
            ws.cell(row=i, column=2).alignment = Alignment2
            ws.cell(row=i, column=4).alignment = Alignment2
            ws.cell(row=i, column=3).font = Font2
            ws.cell(row=i, column=2).font = Font2
            ws.cell(row=i, column=4).font = Font2


        for i in range(5, len(headers) + 1):
            ws.cell(row=1, column=i).alignment = rotate
            ws.cell(row=1, column=i).fill = orangefill
            ws.column_dimensions[get_column_letter(i + 9)].width = 12

        for i in range(1, len(rc_param) + 1):
            ws2.cell(row=1, column=i).fill = bluefill
            ws2.column_dimensions[get_column_letter(i + 1)].width = 25

        for i in range(1, len(steel_param) + 1):
            ws3.cell(row=1, column=i).fill = bluefill
            ws3.column_dimensions[get_column_letter(i + 1)].width = 25

        ws.insert_cols(5,9)
        ws["E1"].value = "TOTAL QUANTITIES"
        ws["E1"].border = Border2
        ws["E1"].alignment = Alignment2
        ws["E1"].font = Font1
        ws.column_dimensions["E"].width = 15
        ws["F1"].value = "Unit Material Price"
        ws["F1"].border = Border2
        ws["F1"].alignment = Alignment2
        ws["F1"].font = Font1
        ws.column_dimensions["F"].width = 15
        ws["G1"].value = "Unit Construction Price"
        ws["G1"].border = Border2
        ws["G1"].alignment = Alignment2
        ws["G1"].font = Font1
        ws.column_dimensions["G"].width = 15
        ws["H1"].value = "Unit Material + Construction Price"
        ws["H1"].border = Border2
        ws["H1"].alignment = Alignment2
        ws["H1"].font = Font1
        ws.column_dimensions["H"].width = 15
        ws["I1"].value = "Total Material Price"
        ws["I1"].border = Border2
        ws["I1"].alignment = Alignment2
        ws["I1"].font = Font1
        ws.column_dimensions["I"].width = 15
        ws["J1"].value = "Total Construction Price"
        ws["J1"].border = Border2
        ws["J1"].alignment = Alignment2
        ws["J1"].font = Font1
        ws.column_dimensions["J"].width = 15
        ws["K1"].value = "Total Price"
        ws["K1"].border = Border2
        ws["K1"].alignment = Alignment2
        ws["K1"].font = Font1
        ws.column_dimensions["K"].width = 15
        ws["L1"].value = "Unit Manhour"
        ws["L1"].border = Border2
        ws["L1"].alignment = Alignment2
        ws["L1"].font = Font1
        ws.column_dimensions["L"].width = 15
        ws["M1"].value = "Total Manhour"
        ws["M1"].border = Border2
        ws["M1"].alignment = Alignment2
        ws["M1"].font = Font1
        ws.column_dimensions["M"].width = 15
        ws.sheet_view.zoomScale = 70

        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)+9):
            for i in range(0, len(headers)+9):
                cell[i].border = Border1

        for cell in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(rc_param)):
            for i in range(0, len(rc_param)):
                cell[i].border = Border1
                cell[i].alignment = Alignment2

        for cell in ws3.iter_rows(min_row=2, max_row=ws3.max_row, min_col=1, max_col=len(steel_param)):
            for i in range(0, len(steel_param)):
                cell[i].border = Border1
                cell[i].alignment = Alignment2


        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers) + 9):
            if cell[3].value == None:
                for i in range(0, len(headers) + 9):
                    cell[i].fill = grayfill

        last_cell = str(ws.cell(row=1, column=len(headers) + 9)).split(sep=".")[1]
        last_cell = ''.join(filter(str.isalpha, last_cell))
        print(last_cell)
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=5).border = Border1
            ws.cell(row=i, column=5).alignment = Alignment2
            if ws.cell(row=i, column=4).value != None:
                ws.cell(row=i, column=5).value = "=SUM(N{}:{}{})".format(i, last_cell, i)
                ws.cell(row=i, column=8).value = "=(F{}+G{})".format(i, i)
                ws.cell(row=i, column=9).value = "=(F{}*E{})".format(i,  i)
                ws.cell(row=i, column=10).value = "=(G{}*E{})".format(i, i)
                ws.cell(row=i, column=11).value = "=((G{}+F{})*E{})".format(i, i ,i)
                ws.cell(row=i, column=13).value = "=(L{}*E{})".format(i, i)
                ws.cell(row=i, column=6).fill = bluefill
                ws.cell(row=i, column=7).fill = bluefill
                ws.cell(row=i, column=12).fill = bluefill
            if ws.cell(row=i, column=3).value == "C25/30 Concrete for Road works":
                for j in range(14,len(headers) + 10):
                    ws.cell(row=i, column=j).value = ""



        wb.save(filename_output)

    def open_output(self):
        cur_dir = os.getcwd()
        path = cur_dir + "/" + filename_output
        os.startfile(path)

if __name__ == '__main__':
    ps = pricesheet()
    ps.print()