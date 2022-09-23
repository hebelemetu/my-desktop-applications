from dbase import Database
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import  PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from xlsxwriter.workbook import Workbook
import xlsxwriter.exceptions
import pandas as pd
import os
import tkinter as tk


class pricesheet:
    def __init__(self):

        self.db = Database('Database/foundation.db')

    def table_output(self):
        self.db.drop_table_works()
        conn = sqlite3.connect('Database/foundation.db')
        cur = conn.cursor()

        dfs = pd.read_excel('Price_Sheet/Price_Sheet.xlsx', sheet_name=None,engine='openpyxl')
        for table, df in dfs.items():
            df.to_sql(table, conn)

    def print(self):
        global filename_output
        for version in range(0, 1500):
            if os.path.isfile('Output/GT_FOUNDATION_TOOL_OUTPUT{}.xlsx'.format(version)):
                print("File exist")
            else:
                print("File not exist")
                filename_output = 'Output/GT_FOUNDATION_TOOL_OUTPUT{}.xlsx'.format(version)
                break

        workbook = Workbook(filename_output)
        worksheet = workbook.add_worksheet('BOQ')
        wrap_format = workbook.add_format(
            {'bold': True, 'font_color': 'black', 'align': 'center', 'border': 2, 'text_wrap': True,
             'valign': 'vcenter'})

        conn = sqlite3.connect('Database/foundation.db')
        c = conn.cursor()

        c.execute("select * from works")
        mysel = c.execute("select * from works")
        headers = []
        for k, desc in enumerate(mysel.description):
            worksheet.write(0, k, desc[0], wrap_format)
            headers.append(desc[0])

        types = [row[1] for row in self.db.fetchResults()]
        for k, type in enumerate(types):
            worksheet.write(1, k+4, type, wrap_format)

        for i, row in enumerate(mysel):
            for j, value in enumerate(row):
                worksheet.write(i + 2, j, value)
        print(headers)

        while True:
            try:
                workbook.close()
            except xlsxwriter.exceptions.FileCreateError as e:
                MsgBox = tk.messagebox.showwarning('WARNING', 'PLEASE CLOSE {}'.format(filename_output))
                if MsgBox == 'Ok':
                    continue

            break
        # MODIFY
        wb = load_workbook(filename_output)
        ws = wb[wb.sheetnames[0]]

        Border1 = Border(left=Side(border_style='thin'), bottom=Side(border_style='thin'), top=Side(border_style='thin'),
                         right=Side(border_style='thin'))
        Border2 = Border(left=Side(border_style='thick'), bottom=Side(border_style='thick'),
                         top=Side(border_style='thick'),
                         right=Side(border_style='thick'))
        Font1 = Font(name='Times New Roman', size=11, bold=True)
        Font2 = Font(name='Times New Roman', size=11, bold=False)
        Alignment1 = Alignment(vertical="center", wrap_text=True, horizontal="left")
        Alignment2 = Alignment(vertical="center", wrap_text=True, horizontal="center")
        rotate = Alignment(textRotation=90, vertical="center", wrap_text=True, horizontal="center")
        grayfill = PatternFill(start_color='00C0C0C0',
                               end_color='00C0C0C0',
                               fill_type='solid')

        orangefill = PatternFill(start_color='00FF9900',
                                 end_color='00FF9900',
                                 fill_type='solid')

        bluefill = PatternFill(start_color='d6e5f2',
                                 end_color='d6e5f2',
                                 fill_type='solid')

        columns = ["C"]

        for column in columns:
            ws.column_dimensions[column].width = 40

        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=3).alignment = Alignment1
            ws.cell(row=i, column=2).alignment = Alignment2
            ws.cell(row=i, column=4).alignment = Alignment2
            ws.cell(row=i, column=3).font = Font2
            ws.cell(row=i, column=2).font = Font2
            ws.cell(row=i, column=4).font = Font2


        for i in range(5, len(headers) + 1):
            ws.cell(row=1, column=i).alignment = rotate
            ws.cell(row=1, column=i).fill = orangefill
            ws.column_dimensions[get_column_letter(i + 9)].width = 12


        ws.insert_cols(5,9)
        ws["E1"].value = "TOTAL QUANTITIES"
        ws["E1"].border = Border2
        ws["E1"].alignment = Alignment2
        ws["E1"].font = Font1
        ws.column_dimensions["E"].width = 15
        ws["F1"].value = "Unit Material Price"
        ws["F1"].border = Border2
        ws["F1"].alignment = Alignment2
        ws["F1"].font = Font1
        ws.column_dimensions["F"].width = 15
        ws["G1"].value = "Unit Construction Price"
        ws["G1"].border = Border2
        ws["G1"].alignment = Alignment2
        ws["G1"].font = Font1
        ws.column_dimensions["G"].width = 15
        ws["H1"].value = "Unit Material + Construction Price"
        ws["H1"].border = Border2
        ws["H1"].alignment = Alignment2
        ws["H1"].font = Font1
        ws.column_dimensions["H"].width = 15
        ws["I1"].value = "Total Material Price"
        ws["I1"].border = Border2
        ws["I1"].alignment = Alignment2
        ws["I1"].font = Font1
        ws.column_dimensions["I"].width = 15
        ws["J1"].value = "Total Construction Price"
        ws["J1"].border = Border2
        ws["J1"].alignment = Alignment2
        ws["J1"].font = Font1
        ws.column_dimensions["J"].width = 15
        ws["K1"].value = "Total Price"
        ws["K1"].border = Border2
        ws["K1"].alignment = Alignment2
        ws["K1"].font = Font1
        ws.column_dimensions["K"].width = 15
        ws["L1"].value = "Unit Manhour"
        ws["L1"].border = Border2
        ws["L1"].alignment = Alignment2
        ws["L1"].font = Font1
        ws.column_dimensions["L"].width = 15
        ws["M1"].value = "Total Manhour"
        ws["M1"].border = Border2
        ws["M1"].alignment = Alignment2
        ws["M1"].font = Font1
        ws.column_dimensions["M"].width = 15
        ws.sheet_view.zoomScale = 70

        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)+9):
            for i in range(0, len(headers)+9):
                cell[i].border = Border1


        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers) + 9):
            if cell[3].value == None:
                for i in range(0, len(headers) + 9):
                    cell[i].fill = grayfill

        last_cell = str(ws.cell(row=1, column=len(headers) + 9)).split(sep=".")[1]
        last_cell = ''.join(filter(str.isalpha, last_cell))
        print(last_cell)
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=5).border = Border1
            ws.cell(row=i, column=5).alignment = Alignment2
            if ws.cell(row=i, column=4).value != None:
                ws.cell(row=i, column=5).value = "=SUM(N{}:{}{})".format(i, last_cell, i)
                ws.cell(row=i, column=8).value = "=(F{}+G{})".format(i, i)
                ws.cell(row=i, column=9).value = "=(F{}*E{})".format(i,  i)
                ws.cell(row=i, column=10).value = "=(G{}*E{})".format(i, i)
                ws.cell(row=i, column=11).value = "=((G{}+F{})*E{})".format(i, i ,i)
                ws.cell(row=i, column=13).value = "=(L{}*E{})".format(i, i)
                ws.cell(row=i, column=6).fill = bluefill
                ws.cell(row=i, column=7).fill = bluefill
                ws.cell(row=i, column=12).fill = bluefill
            if ws.cell(row=i, column=3).value == "C25/30 Concrete for Road works":
                for j in range(14,len(headers) + 10):
                    ws.cell(row=i, column=j).value = ""



        wb.save(filename_output)

    def open_output(self):
        cur_dir = os.getcwd()
        path = cur_dir + "/" + filename_output
        os.startfile(path)

if __name__ == '__main__':
    ps = pricesheet()
    ps.print()